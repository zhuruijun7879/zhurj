import json
import time
import random
import logging
from pathlib import Path

import requests
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ================= 配置 =================
BASE_DIR = Path(__file__).parent / "x"
URL = "http://elearning.js.sgcc.com.cn/sdext-selftest-api/exam/testpaper/question"
PAGE_CODE = "1781609675403"

REQUEST_INTERVAL = 0.8        # 基础节流
RANDOM_DELAY = 0.6            # 随机抖动（防风控）
MAX_RETRY = 3                 # 单题最大重试
TIMEOUT = 10


# ================= 日志系统 =================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    handlers=[
        logging.FileHandler("run.log", encoding="utf-8"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("excel-batch")


# ================= HTTP Session（带重试） =================
def create_session():
    session = requests.Session()

    retry = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=[429, 500, 502, 503, 504],
        allowed_methods=["POST"]
    )

    adapter = HTTPAdapter(max_retries=retry)
    session.mount("http://", adapter)
    session.mount("https://", adapter)

    return session


# ================= 请求头 =================
def build_headers(auth, cookie):
    return {
        "Accept": "application/json, text/plain, */*",
        "Content-Type": "application/json;charset=UTF-8",
        "Authorization": auth,
        "Cookie": cookie,
        "User-Agent": "Mozilla/5.0",
        "Referer": "http://elearning.js.sgcc.com.cn/sdext-online-exam-web/"
    }


# ================= 防封节流 =================
def throttle():
    delay = REQUEST_INTERVAL + random.uniform(0, RANDOM_DELAY)
    time.sleep(delay)


# ================= 安全解析 =================
def safe_extract(data):
    if not isinstance(data, dict):
        return ""

    return (
        data.get("answer")
        or data.get("data", {}).get("answer")
        or data.get("result", {}).get("answer")
        or ""
    )


# ================= 单题请求（带重试） =================
def request_question(session, headers, qid):
    payload = {
        "results": [{"questionId": qid, "userAnswers": [""]}],
        "pageCode": PAGE_CODE
    }

    for attempt in range(1, MAX_RETRY + 1):
        try:
            resp = session.post(URL, headers=headers, json=payload, timeout=TIMEOUT)
            resp.raise_for_status()

            data = resp.json()
            return {"success": True, "data": data}

        except Exception as e:
            logger.warning(f"题目 {qid} 第 {attempt} 次失败：{e}")
            time.sleep(0.5 * attempt)  # 递增退避

    return {"success": False, "error": "重试失败"}


# ================= Excel 处理（支持断点续跑） =================
def process_excel(file_path, session, headers):
    logger.info(f"开始处理文件：{file_path}")

    wb = load_workbook(file_path)
    ws = wb.active

    # 表头
    ws.cell(1, 2, "答案")
    ws.cell(1, 3, "状态")

    for row in range(2, ws.max_row + 1):
        qid = ws.cell(row, 1).value
        old_value = ws.cell(row, 2).value

        # ===== 断点续跑核心 =====
        if old_value:
            ws.cell(row, 3, "已跳过（已有结果）")
            continue

        if not qid:
            ws.cell(row, 3, "空题号")
            continue

        logger.info(f"处理题目 {row-1}: {qid}")

        result = request_question(session, headers, qid)

        if result["success"]:
            answer = safe_extract(result["data"])
            ws.cell(row, 2, str(answer))
            ws.cell(row, 3, "成功")
        else:
            ws.cell(row, 3, result["error"])

        throttle()

    wb.save(file_path)
    logger.info(f"完成并保存：{file_path}")


# ================= 主程序 =================
def main():
    if not BASE_DIR.exists():
        logger.error(f"目录不存在：{BASE_DIR}")
        return

    files = [
        p for p in BASE_DIR.iterdir()
        if p.suffix in [".xlsx", ".xlsm"] and not p.name.startswith("~$")
    ]

    if not files:
        logger.warning("没有找到 Excel 文件")
        return

    auth = input("Authorization：").strip()
    cookie = input("Cookie：").strip()

    headers = build_headers(auth, cookie)

    session = create_session()

    for f in files:
        try:
            process_excel(f, session, headers)
        except Exception as e:
            logger.error(f"文件处理失败 {f}：{e}")

    logger.info("全部任务完成")


if __name__ == "__main__":
    main()